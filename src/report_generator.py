"""
PDF and Excel attendance report generation.
PDF  : Portrait A4, NPS logo header, per-section summary table, colour-coded rows.
Excel: Two sheets — Attendance (formatted) + Summary.
"""
from datetime import datetime
from pathlib import Path
import base64, io

import pandas as pd
import config
import database as db

# ── Logo (decoded once at import) ─────────────────────────────────────────────
_LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAAEAAAABACAYAAACqaXHeAAAYPklEQVR42u2beXxV1bn3v2vtvc88JCcTSQhJCCgKBRVkUDRarW2lam9raG2r1murfe/b6tWqvXYK8Vpaq51926tW69BWS+o8VBHBVAVFAogYUJAQSAIJGU/OvPde6/3jBATElsFr+7nX5/PJJzk7e1i/33rW83ueZ68DH9qH9qH9bzZx5NfrfwYIH/QgGmVDQ4PxTzONDQ0GNMoPyAMaJTQpgIsvbvQNUeD7R2IvYChzzz1Nmf3H9t9EgBYg9Jz6qycJI/S9rOOe5LpuKH/8H+L+2pRm0rLEy1Jn//OlZT9+41BJOOiBNzY2yqamBXpO/TWTXeFvSWeIWabC7/eg34dgcqimNQghSKdz5BwIeOUwMle/ctlN6xobG0VT08GRYB7sA5tGGXfF9T9Lp0Xs5DlV2Su+9klPYUEo/x+RH9TeTAhEPkxqjRQCpfS+TI2eL/b6qDUYcq9z97qn4J1naK1BQ3wkxa9uezrX8lJ7NOA1fgGc1tR08PMhDmXd13+8sWY4kXvL5zXMJ5qvpSgWEf8MMXBoeIRPnX+zSqZsHQ15J7Usbtp8sEvhoDygvh7Z0oJyhTs3l8M6aWa1WxSLGKvWbKY4FiaVyuEPeOjqGiBWGEZrRc62KS0pwOM12blziMnHVPHSy28SifgxpQShGRhMUlgQJBwOIAX0Dybo7Orn42cex5rX2vF5TYaGU1iWQVlJAVIKtmztobKiiL6+OH6/l3DYx8S6Ck44rlYtadlkKu3UA5t3j/l9WwIA2hb1jquYccJ4rYEbf/wIM46vJZWxKS0Js6p1C1WVRUSjQbq6+vnMeTNRWvPwoyu59aeXcsfdz2EakprqUhDw5ptdlJREOaN+Mlnb4dnn1nPKyUeRSedY2rKebNbG47GYPKmSu+5rIeC3iEaCpFI5fH6LxEiWSUeVM7GuguOnVbN46Vu4LqcCdx4spoPSzpaWJldrRM52T7RMwbQp1VJoqKqM0bGtj0QizbZtuygtDhMJ++jdFScU9vPa+q20rtlCOptjV98Q42vH8K2rz2NgMAkaQiEfhiHZ+FY3PT3DhMM+BgaSuK6isCBET28c11EMDCUpKAgSCvpxlcLnNdFaEQx6ae/oo68vzozj66RlgePo6Y2NjbKlpcl9n2JAfi19/OPfLt854myOhoOBRx64SkejQdHV3UcqlSMY9JLN2himgWka9PQMU1YWpXvHAJGwn0g4QCjgJZ3OUVYW47V1WygoCJBKZSmIBtjZM0ww6CEY8LGts59ZJ05k41s7CAa8DI+kMQxJUWEIaUg6u/oZUxqlfzBBJOxnV98IteNK8PktPe/8m8XQYCpTVhY6avGj39t+MHHg7y6BhobJorkZ0tqY5DoqMLYypgoLQhIEVZWlB7xmbHnxPr93WyQCCpg2dfw+x8vLS/b8XVlZigaOnVR1wHuXFkcBqCgvAqC6qnRUFrWoripWff2dvnTWOQbYvnvsR0RAb+8bYlTSpjiuoK6uRIGQm3s6GckmMaTco8vvyKHeo4tKa+pKqwgld2Knh5DSxN5L4vaXT61crGABaW8FHdt6EULs5aoCrXX+GvJ5gOsoAgEP42vHMLGuTK1s3S6Vcj8CLN499vclCNqOnqK1YuL4CgC++scm1na9SdDrR+kDe5kUgkQmxeNX3MbcFxcy8trDGP4YaOc9FqSBmx4mOu0s3jzul1x40c/w+nx5Qg90uhCkMzmOmlDGY3+6jgnjx6A0aFdMPlhc5t8PgHkpcZWaKCVMrCkTOZ2jZ6SfgMeLKSX6PTJhKQSWYeRnUUiEtBDSeO/abfQchIEQYBgGhiH/JgE+n4eBgSS5XI668aXCNDS2cifuPfYjUQEBTaqxsdF0XHec12swtrJI9IwMEM+kMKSB1jqfwf2Nn31Tv3eh2F9sDyEd1hhSkEhl6dkVp6qyWPi8Jsp2qxoaGj2jAVAcAQH5wbzS5il0HF0aCnooLY6KHUN9ZJwsUogjLMQFuM4BSDh4k1KQzdr09A5TXBQRoZAXx9UlcccqOhhC/zYBjQsEQG4kV+IqFYpEvHj8pugdHsBxHYSQR4BdgHYhEAPXPoLbCFxX0dc3gmkaIhr14boEcrlM6d4YDouAhrY2AaBQJVpLURANKoBdiSEU6vArQCEhm0AUjceccg7aVwCm57AbO0rl02iAgmhQaSTKMUr2xnBYBPT2HpuXQC2LlBYURP0aYCAZP/wm1OjM60AR5pRzMSecjnXSVyGXBHG4TSbN8HAKgMJoQCklQFC8N4YjkkFHiEKtIRzJEzCSTR5+A0BIdC6F59QrsY6fj3YdTI8f1xtG25lDcv131EGQSGYBCEcCaK1xlS5832oBlA5rLQgGPACkc5m/EdbE3+mhKhACq2YOZmENhr8QbfqQE+ohl8zL5EGAtx13j0MJoUmn8wSEAvkGDUKED5mAxkakPsDcSiH8oPF68gTkHBuBGO1ViD2aL4XA1S4CMKWxJ4vbR1WEAARKKyRgBoswimoRoTGHFPVLikJIKXCVRghBNpdPrnxeEzQorf3voWviPQloakIJ0Lpx3+NKKRMNhpk/7CoXMQrY0S6GNMg4ORLZNAGPD0e59I4M5Ik6kMSlBsgsuQlnuAs3NYC9/jGctc1g+vIe8jfAK6UJBb389tav8M1vfJLh4RRSCpxRjzCM0dTcVXJ/4HoRxv59fLk3K9t/fupEEIgmlF6EUVraJnc3H/fO96UQZEcBR7xBBlLDjC0Yw6JLb+bFq+/mkct+wS/Ov44pFRPI2rl3SNAaLSTWp36IdfQZ5DpXke1aDdkUxsyL0N4QKPeA8WU3+F274sz/7Cxqqsfw+fPncvm/nk48nkFKsU9Nsg/4UeBiPm777y72vTsINmLQhBOMb7s686Nxk9Mi8u9i/uurodnVizDqfytsBNi2s+chJaEYd36hkZriCp5pW8GM6mOZPu4YtuzqYtrYown5/Px4yd1YhvkO59JAZBMQKsGa8FF0egi0izj2bJyty3GycThAqiyEIJXO4fNafPvac/nMOdNZ90Y7UyfXcs0V59K6pp1MJj8223ZAgJSGrRswLt+yQ4j5ONsWXeWPtj+2wO5abgLf1IswxHzc/AzvyLsHSqW8Hn2KzA69HF84/tYtN59eLebjplxzUI4OAsBjeagqKKUqVk5pKEZdyVjGRIr566bVVBaWIiX8+59vpjcxgGWY6N2ItAJPAHvR18iu/iMIgZAW2dceJHP/vyJM/7vyfkMKUqkcn/vMTJ555Hq+dulZlJYWUVYSpXXtJn7yq8dZ39ZJYUEgv7rS2bwDCZUQzbi3r7rNGfnR0V+IvvnnVRGffZ1EhxsbkbyRPy0f9AZRYj6uxkiQ065tu0ZI2P+3KL11jf5xyTUTo/0eV2udiKcFQHm0hFe3vcEZv7qMl7e+ju06LG9fy6SyGkJeP+s6N7G2cyOF/giuct+dB1g+DMODt/I4PJXTME0vwvTmE6T9ujVKa0xT8NVLzsRxHHb2DALw6uq3+fLXbueue1tIJHPs7k4Px7MIIfAYxrD+edXc4YUT/+rX2T9oxz2WlHaFa3c3NaGIYWiNkAJ065mXyeGbppztMfW8tK2k1yMzjmunw6ZdiDZvPrF0x5iso/TgYMIAKA0XEvTmm5sX3vMdFjz5G/pHBqksLGUkm+KGp27fUyLrA5UXQuLuXI/WCrTC3fH6KPj9Zt+QDMdTnHHaZEIBi9a1b1MQDbK1o4frFyzCMEyi0SCmKSkpzqve4GDcTNqu+nTVm5PdlPOCxx6Zm03HXdM0yLpaSqGn9v50erm4kqwQaHPnwilzgrueuSfk1RPjjsJnuu7TO2ofXjNQ5Hz/+Lappifx2g9br7q1wv/gd3r7R0BDdVE5UhhIIfF7vOyI9/HQa0vZ1NfJ6m0beLO3naAn8B59AoWQFmqgY/QTuAPtCMN6VwSzHZdoJMDnPjOLZCrLSbMmsb5tG3f9vgXHUfh9BrbtYBiSyooYALv6Roh4kct6j7n/itrFs40xlXPW6FOUNbRpw/jM6ilC+v/FP9J75vCN45/L+IsXmG7W+5pjZb+VzOaujwTlibmsbL9xVXVVZ786+d7WYzdPrfF3Tit9PDwYN3cNDuVKhgZTuraoQgQtP47rIIXEY1ps7N3K6u0b8VvePeDlgSRQazAs1KgEYpjo4S4w9q0FpBRkMjbnzDuB+EiK/3f7El5YvpGe3jimaRAK+nDdPMEBv4fqcUUMx5N6YDAjfF5z8NHHbmy9u+5Tj27sHn/SohXDZrUn0Pncp8yYEFSkHO0R6A7Dpt+sbGpNAQ/fdtttT1yavOXap3srj+0acM752OxaI1IYO3rx0lXf1eq1qwKBQplIBNmweZuYM3MSdeVjWd+1maDhRymFR1r4gl60Vu/ZIXrHt01IDaISPSAtdHIgHwP2yT00Pp/F2te3s+T5NgYHkwSDPkIhXz7RUfnWWDbrUDW2iOqqUja+1c3gcAJU3KmqPf35m573z66pUhxX7RMvrTNObB0qbZ1emlgzbJVdU3nNio3QnpdB3YAhLr/cvRwWHjN95h9CZjLyjSsudNrbO42jJ45V69s6gm+0baK3dzv/54omzvv4qYR0ECNnIKIajUYrjXY1B+zB6HfHAOwUanA7GFa+EPL437UEhBB0dw9hGIJYYQilVP6VGSCERBoSrV3qaov464ut3PqbB8Vw/0amfaSmZPJZn6ovjgW4+EvzWLx0pbtm/e+Krlg+ufPlF++/FLagF2EwH2UCiOZGLWhSHzvrgpmbt8U/f8rc43LRSNBcsmyV+PZ1XzZmzNilt249Xvj9Hu79w1M0P/ocmWSOYCCEjrlYZaBjCjei0F6FMOU77SAHhCFGS4rRF8wItNLYu94GaeEqhdAyHxC1wNXskUOPx0RrUFoghDkaK11sO42bSTEy3Mfzy7ax+JknmTK5jjt+/S0SybSefMx49atf/8no2N7D2Z84WTY/uNRt29D+L2ef/YUbnjzxj9vF/EYNTdrMt77bRHMz9MeZ6fX6GBxMexYvWUl1ValrGJJEMi1rasoxpOCSi84hPpKksqKEhx5+njfbOhjYFMdWNq7XRYU1KSuNiirMIgtTGjg5F2EKwl6Nx+ciJShXI0c2gDQJBUD4nfzywcXwCbzefEzQ2sGxc9h2FqWyOHYSpWyiYS+WJfj0F+pJpXJMmzqB0pIY3Tv6aW/vFH6f15g5Y7J+ZskranDRs2JoKGkEguHCnX3ZL4omFtbXP2+2tKDEfrKr5869aOrO/qGvaK0/VzamrHTixCom1lUy5dhaRwghbduVhbEo99z7OLNmTqauroqO7d3c8vNFaG3g1Q611ZXkcjm2dfeQwyEaCVPMIDE5TCyg8YgshV4Xb6wSQwrsvq3ktIUjLAYTOVRBNf2BY9nY1o6rFV6fQawgjBBw1FHVZNI5HFdxycXn4doOLy5fwxkfnUWsMMzd9z3JMZNq1NLnX1VHH1VjNj+0jGQyjWXK14MB69fY4s+trff377Xh4cB2wQVXF69p2/qZXMb+omWZc4uLi+TUqRMojAaZNXMKff3DaKWoP/UE/rLkVe765W+YUO7jjcEID9zbSMsLa6gdV8GKl9cTjQToHUrT3TNI7644Y6vK2NC2FSkVrlIUxgoI+n0Mx5PUVJez/vWNdHd2o6SH4yaVc83VF7FhwxbOOnMOW97uZCSV4me/up9bFl7JyEiCQMBH+9ZutrR3u2++1SHfertTpNMOw8ODucKC0OJQyHvnWafVPdHU1OQcREOkUTY0tIn77/9pH3C7lNx+0mkXTe3rG/j0M88uP0+56oSVq15XpaUxGY1ESCZTrN3YSVlIM8aMM1gxCUNI3t7UyZn1s5hQV0X/QBy/zyQSDtLRsQNDCnLnzKAgGiGVyRArjNDT08+69Zu58IJ53HHvkzz8hweYUZHgjWQJZSUFPPb4dj52xizKK4oRO/v48pfOZn3bZtrbuzFNg6efXa6SKduwczamIdf5vNaDU4+uaH7qqTs3AKxogYaGBqO5uVntHZYPQECTGn2dJOrr642Wlhb3xaX3rgPWSSlumF1/5Td29tm/fLt9kxsKmMbiJa+Qy+XALGV1n2LuDOjq6iUWC9DX309hQZgdO3opK4vxX3c8zLxPzCWVyuC4LsdNm8TgUILWtRsJBQP0DyTI2g6mZaGUxjBM+nuHSCaSjK8Zw0OPPIttazZt7iRn2yx/+XWE0GiFCkeKpMf03lEYCdx5xb99ctX8+fPdvSe0ublZNTc3u4e9QeKd9+03qNkf/d5LqbQ+6aQTy92LvzDHeGH567yx+kk6tg/Q1R/D1RaBYAiPJYgVBCkqDmKgWLxkBddfcyElxQU89tRfafjsRxkaHGHDhnZOmXs8r6xcj9fnJZXO8tDjy3GlH0vblJSWYtua7V39CHeEaBjGlJdz2qknsGLlNrdvSBgeU7668vkfzNR79jTUmy0tp6m/93L0kDp7u13opNOvPcrR3tbheMa/4PpzxefPP0XozFria89hx44Em3YWs2m7l46dIbbv8rBr0MtwyouWYQaHc7gKhLBwXReQGFIgpEQIhes4WJZLYcQi6MsRDTpEA3HGlWWZVDVEXbWf6tl3UVN3Ag899rJesPBR7fN7sl4pZry4tGnj9OmXGa2ttx90n/2QW5sNDYuM5ub57uzTr5uvtP9Pg4PDzi0/+Jxx7rzZwkm0Ym6bB7oHLAGuBkfgpC1sF+JJk0TGw0jaSzLrJWuD0BpXgWXlGy2xUArT0Pg9NpGgTcifwwwIEA5YMah+DDwns/i51frr37zPjRYUmIaRvvCVpTf9fvfYDgXPIfeh29qadX19o/nS8wtfH1s72/X6omc+vWSdc1RdkTHh6OmkvZ9E9S+D+C6060E7GjPo8NrmGvpHomQdH9LwkMhEmFQVp3ugnGNrB+jorWDW0TvYuquMY6r7KC5O4rMcpDTQaYecmEiu8i9YwRksa1nLVdc/4ITDUcuQ6RteWXbTL+vrG82nnvq6e6h4DqsR39HRovIk/OD5qtrZBYYVPPnJp9fY4yojcvKU44UuPB8y6zDSb6GFgfDC4pXH8OiKqXT3R0jnTAYSPla0jcPvs+nojfL06kmE/Rl+v3QWY4sHqCiLo7ImOA4qeiZMfBRfaAJ/Wfwq3/zuItvnC1mWTN/68rIfXVdf32i2tDQ5h4PlsLe7dnS06IaGRcbiJ6/8S1XtSQWmFTj5iadXK69XixNnTBOy+EJspZCJF5BC0d1fSn8yyITyPqTQZHMWlUVxNnaOIRrIUFM6QP9IEEMKIgGbSeO6cG1wy7+FWXcPphnmd/ct0d9f+KgKBMOmJTO3rlj2w280NCwynnrq6+pwcRzxZumGhgbZ3Nzszjnj+u+6yvefgwNDnHfONPd71zUYkXCA7FALRtc1pPpXIQw/hvSQykoyOYPSgiTb+6KURUcwDMVIOkRBYJBs2iFQchxOxU/wFn6URDLDD37c7D746BqjMFaAQaZpxbIfLqChwWA/Xf+gCRglYZFsbp7vzjn9Py7QwnP7wFAmVFcTdb5z7XnGybMnC3Cxd/wEq/dmyPXlsw9DgANYGlwj3w3WoGQRaszVmOXXAhYvr9zAjTc/5mzaMmDGor6UIPe1Fct+dN9owDsi8O8XAaO6m1+Hs8/81kek9v42mWGmnUsx/9MnuP922TyjpDiCsndgd/0Mc+BOhD2AMEYrYAXKiuEUXoKn8iqkp5K+/gT/9du/uA88uFKaVkAE/KJVOMmvLG+5Ze2RrPn/NgL2JqGhocHTOTDp+67LdUPxnDWmxOde8qW5Yv5nT5cBv4F2Osl134HRfyegcGOX4Km4HGGNI51RND+0TN1134t6R2/KiEa8jmXqWyx7e1NLyz2Z9xP8+07A6Ps1yehG5bmn/8d0LT0L0zl9VmIkTV1tofvFz80R5847RYZDJjA0elEBiaTLE0++oH6/aIV+a8uAEQ4G8Ht5Tmrn2y8sXbgyf+tGebCboP9xBLC7jmg0ds/UKWd877MO4vp0Rk1PpbKMGxtW5807Xn/iYzMFQrD42Vf1I0+uER3bh6Xf7yPgE2ulVD98acmNi/ZKvo54vX+QBOzlDQs0CN3Q0GDsGDxqvquMK9MZd1Yq4xAKGgggkXTx+0x8PuNV0xC/PCPHA00tTQ5o0di4QLzfs/7BEbBvDbEnS5t7VuMntOLL2WxuthYIr+V5xbS4+4Wnm57S+6Xc/A8y0dCwaJ/E6+KLG30XX9zo27/W4IP//sUHaw0NDcY+X7ra//P/IhP/42f7Q/vQPrR/avv/b1ZKO+fRgzkAAAAASUVORK5CYII="
_LOGO_BYTES = base64.b64decode(_LOGO_B64)


# ══════════════════════════════════════════════════════════════════════════════
# PDF
# ══════════════════════════════════════════════════════════════════════════════

def generate_pdf_report(target_date: str, class_name: str = None,
                        stream: str = None, section: str = None) -> str:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.units import cm, mm
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from reportlab.platypus import (
        SimpleDocTemplate, Table, TableStyle, Paragraph,
        Spacer, HRFlowable, Image, KeepTogether,
    )
    from reportlab.lib.utils import ImageReader

    # ── File path ──────────────────────────────────────────────────────────────
    fname = f"attendance_{target_date}"
    if class_name: fname += f"_{class_name}"
    if stream:     fname += f"_{stream}"
    if section:    fname += f"_Sec{section}"
    path = config.REPORT_DIR / (fname + ".pdf")

    doc = SimpleDocTemplate(
        str(path), pagesize=A4,
        leftMargin=1.8*cm, rightMargin=1.8*cm,
        topMargin=1.5*cm, bottomMargin=2*cm,
    )

    # ── Colour palette (NPS) ──────────────────────────────────────────────────
    NPS_BLUE    = colors.HexColor("#1d4ed8")
    NPS_BLUE_LT = colors.HexColor("#dbeafe")
    NPS_ORANGE  = colors.HexColor("#f97316")
    NPS_GREEN   = colors.HexColor("#16a34a")
    NPS_GREEN_LT= colors.HexColor("#dcfce7")
    AMBER_LT    = colors.HexColor("#fef3c7")
    RED_LT      = colors.HexColor("#fee2e2")
    GREY_LIGHT  = colors.HexColor("#f8fafc")
    GREY_LINE   = colors.HexColor("#e2e8f0")
    DARK        = colors.HexColor("#0f172a")
    MID         = colors.HexColor("#334155")
    MUTED       = colors.HexColor("#94a3b8")

    # ── Styles ─────────────────────────────────────────────────────────────────
    def style(name, **kw):
        return ParagraphStyle(name, **kw)

    S_SCHOOL   = style("school",   fontSize=13, fontName="Helvetica-Bold",
                       textColor=DARK, alignment=TA_LEFT, leading=16)
    S_SUBTITLE = style("subtitle", fontSize=8.5, fontName="Helvetica",
                       textColor=MID, alignment=TA_LEFT, leading=12)
    S_TITLE    = style("title",    fontSize=10, fontName="Helvetica-Bold",
                       textColor=NPS_BLUE, alignment=TA_LEFT, spaceAfter=2)
    S_SECTION_HDR = style("shdr", fontSize=8, fontName="Helvetica-Bold",
                          textColor=MUTED, alignment=TA_LEFT,
                          textTransform="uppercase", letterSpacing=0.8)

    records = db.get_attendance_by_date(target_date, class_name, stream, section)

    # Deduplicate: first detection per student
    seen = {}
    for r in records:
        if r["student_id"] not in seen:
            seen[r["student_id"]] = r
    unique = list(seen.values())

    # ── Build subtitle line ────────────────────────────────────────────────────
    date_fmt = datetime.strptime(target_date, "%Y-%m-%d").strftime("%d %B %Y")
    parts = [date_fmt]
    if class_name: parts.append(f"Class {class_name}")
    if stream:     parts.append(stream)
    if section:    parts.append(f"Section {section}")
    subtitle_str = "  ·  ".join(parts)

    elements = []

    # ── Header: logo + school name + metadata ─────────────────────────────────
    logo_img = Image(io.BytesIO(_LOGO_BYTES), width=1.4*cm, height=1.4*cm)
    logo_img.hAlign = "LEFT"

    header_data = [[
        logo_img,
        [
            Paragraph(config.SCHOOL_NAME, S_SCHOOL),
            Paragraph("Attendance Report  ·  " + subtitle_str, S_SUBTITLE),
        ],
        Paragraph(
            f"Generated<br/>{datetime.now().strftime('%d %b %Y  %H:%M')}",
            style("gen", fontSize=7.5, fontName="Helvetica",
                  textColor=MUTED, alignment=TA_RIGHT, leading=11)
        ),
    ]]
    header_tbl = Table(header_data, colWidths=[1.8*cm, None, 3.5*cm])
    header_tbl.setStyle(TableStyle([
        ("VALIGN",       (0,0), (-1,-1), "MIDDLE"),
        ("LEFTPADDING",  (0,0), (-1,-1), 0),
        ("RIGHTPADDING", (0,0), (-1,-1), 0),
        ("TOPPADDING",   (0,0), (-1,-1), 0),
        ("BOTTOMPADDING",(0,0), (-1,-1), 0),
    ]))
    elements.append(header_tbl)
    elements.append(Spacer(1, 3*mm))
    elements.append(HRFlowable(width="100%", thickness=1.5,
                               color=NPS_BLUE, spaceAfter=4*mm))

    # ── KPI pill row ──────────────────────────────────────────────────────────
    total_enrolled = len(db.list_students(class_name, stream, section))
    present_count  = len(unique)
    absent_count   = total_enrolled - present_count
    pct = round(present_count / total_enrolled * 100, 1) if total_enrolled else 0

    avg_conf = (
        sum(r.get("confidence") or 0 for r in unique) / len(unique)
        if unique else 0
    )

    kpi_data = [[
        _kpi_cell("Present",  str(present_count),  NPS_BLUE),
        _kpi_cell("Absent",   str(absent_count),   NPS_ORANGE),
        _kpi_cell("Rate",     f"{pct}%",           NPS_GREEN),
        _kpi_cell("Avg Conf", f"{avg_conf:.3f}",   colors.HexColor("#7c3aed")),
    ]]
    kpi_tbl = Table(kpi_data, colWidths=["25%", "25%", "25%", "25%"])
    kpi_tbl.setStyle(TableStyle([
        ("LEFTPADDING",  (0,0), (-1,-1), 4),
        ("RIGHTPADDING", (0,0), (-1,-1), 4),
        ("TOPPADDING",   (0,0), (-1,-1), 4),
        ("BOTTOMPADDING",(0,0), (-1,-1), 4),
    ]))
    elements.append(kpi_tbl)
    elements.append(Spacer(1, 5*mm))

    # ── Attendance table ──────────────────────────────────────────────────────
    elements.append(Paragraph("Attendance Records", S_TITLE))
    elements.append(Spacer(1, 1*mm))

    has_stream = any(r.get("stream") for r in unique)
    if has_stream:
        hdrs = ["#", "Student ID", "Name", "Class", "Stream", "Sec", "Roll", "Time", "Confidence"]
        col_w = [0.6*cm, 2.1*cm, 4.5*cm, 1.3*cm, 2*cm, 0.9*cm, 0.9*cm, 1.4*cm, 2*cm]
    else:
        hdrs = ["#", "Student ID", "Name", "Class", "Section", "Roll", "Time", "Confidence"]
        col_w = [0.6*cm, 2.3*cm, 5*cm, 1.5*cm, 1.2*cm, 1*cm, 1.5*cm, 2*cm]

    tbl_data = [hdrs]
    style_cmds = [
        # Header row
        ("BACKGROUND",    (0,0), (-1,0),  NPS_BLUE),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0,0), (-1,0), 7.5),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("VALIGN",        (0,0), (-1,-1), "MIDDLE"),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, GREY_LIGHT]),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",       (0,1), (-1,-1), 7.5),
        ("GRID",          (0,0), (-1,-1), 0.3, GREY_LINE),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("RIGHTPADDING",  (0,0), (-1,-1), 4),
    ]

    sorted_rows = sorted(
        unique,
        key=lambda r: (r.get("class_name",""), r.get("stream",""),
                       r.get("section",""), r.get("roll_no",0))
    )

    for i, r in enumerate(sorted_rows, 1):
        conf = r.get("confidence") or 0
        time_str = str(r.get("detected_at",""))[-8:] if r.get("detected_at") else "—"
        if has_stream:
            row = [str(i), r["student_id"], r["name"], r.get("class_name",""),
                   r.get("stream","—"), r.get("section",""),
                   str(r.get("roll_no","")), time_str, f"{conf:.3f}"]
        else:
            row = [str(i), r["student_id"], r["name"], r.get("class_name",""),
                   r.get("section",""), str(r.get("roll_no","")),
                   time_str, f"{conf:.3f}"]
        tbl_data.append(row)

        # Colour-code confidence column (last col)
        row_idx = i
        if conf >= 0.45:
            style_cmds.append(("BACKGROUND", (-1, row_idx), (-1, row_idx), NPS_GREEN_LT))
            style_cmds.append(("TEXTCOLOR",  (-1, row_idx), (-1, row_idx), NPS_GREEN))
        elif conf >= 0.32:
            style_cmds.append(("BACKGROUND", (-1, row_idx), (-1, row_idx), AMBER_LT))
            style_cmds.append(("TEXTCOLOR",  (-1, row_idx), (-1, row_idx), colors.HexColor("#92400e")))
        else:
            style_cmds.append(("BACKGROUND", (-1, row_idx), (-1, row_idx), RED_LT))
            style_cmds.append(("TEXTCOLOR",  (-1, row_idx), (-1, row_idx), colors.HexColor("#991b1b")))

    if len(tbl_data) == 1:
        tbl_data.append(["—"] * len(hdrs))

    tbl = Table(tbl_data, colWidths=col_w, repeatRows=1)
    tbl.setStyle(TableStyle(style_cmds))
    elements.append(tbl)

    # ── Per-section breakdown summary ─────────────────────────────────────────
    elements.append(Spacer(1, 6*mm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=GREY_LINE))
    elements.append(Spacer(1, 3*mm))
    elements.append(Paragraph("Section Breakdown", S_TITLE))
    elements.append(Spacer(1, 1*mm))

    # Group by class / stream / section
    from collections import defaultdict
    groups = defaultdict(lambda: {"present": 0, "total": 0})
    for r in unique:
        key = (r.get("class_name",""), r.get("stream",""), r.get("section",""))
        groups[key]["present"] += 1

    # Get total enrolled per group
    all_students = db.list_students(class_name, stream, section)
    for s in all_students:
        key = (s["class_name"], s.get("stream",""), s["section"])
        groups[key]["total"] += 1

    sum_hdrs = ["Class", "Stream", "Section", "Present", "Total", "Absent", "Rate"]
    sum_data = [sum_hdrs]
    for (cls, strm, sec), counts in sorted(groups.items()):
        pres  = counts["present"]
        total = counts["total"]
        absent = total - pres
        rate   = f"{round(pres/total*100,1)}%" if total else "—"
        sum_data.append([cls, strm or "—", sec, str(pres), str(total), str(absent), rate])

    sum_style = [
        ("BACKGROUND",    (0,0), (-1,0),  NPS_BLUE),
        ("TEXTCOLOR",     (0,0), (-1,0),  colors.white),
        ("FONTNAME",      (0,0), (-1,0),  "Helvetica-Bold"),
        ("FONTSIZE",       (0,0), (-1,0), 7.5),
        ("ALIGN",         (0,0), (-1,-1), "CENTER"),
        ("FONTNAME",      (0,1), (-1,-1), "Helvetica"),
        ("FONTSIZE",       (0,1), (-1,-1), 7.5),
        ("ROWBACKGROUNDS",(0,1), (-1,-1), [colors.white, GREY_LIGHT]),
        ("GRID",          (0,0), (-1,-1), 0.3, GREY_LINE),
        ("TOPPADDING",    (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
        ("LEFTPADDING",   (0,0), (-1,-1), 4),
        ("RIGHTPADDING",  (0,0), (-1,-1), 4),
    ]

    # Colour-code rate column
    for i, row in enumerate(sum_data[1:], 1):
        try:
            val = float(row[-1].replace("%",""))
            if val >= 75:
                sum_style.append(("BACKGROUND", (-1,i), (-1,i), NPS_GREEN_LT))
                sum_style.append(("TEXTCOLOR",  (-1,i), (-1,i), NPS_GREEN))
                sum_style.append(("FONTNAME",   (-1,i), (-1,i), "Helvetica-Bold"))
            elif val >= 50:
                sum_style.append(("BACKGROUND", (-1,i), (-1,i), AMBER_LT))
                sum_style.append(("TEXTCOLOR",  (-1,i), (-1,i), colors.HexColor("#92400e")))
            else:
                sum_style.append(("BACKGROUND", (-1,i), (-1,i), RED_LT))
                sum_style.append(("TEXTCOLOR",  (-1,i), (-1,i), colors.HexColor("#991b1b")))
        except ValueError:
            pass

    sum_col_w = [2*cm, 2.5*cm, 2*cm, 1.8*cm, 1.8*cm, 1.8*cm, 1.8*cm]
    sum_tbl = Table(sum_data, colWidths=sum_col_w, repeatRows=1)
    sum_tbl.setStyle(TableStyle(sum_style))
    elements.append(sum_tbl)

    # ── Footer note ────────────────────────────────────────────────────────────
    elements.append(Spacer(1, 8*mm))
    elements.append(HRFlowable(width="100%", thickness=0.5, color=GREY_LINE))
    elements.append(Spacer(1, 2*mm))
    elements.append(Paragraph(
        "Confidence legend:  "
        "<font color='#16a34a'>■ ≥ 0.45 High</font>  "
        "<font color='#92400e'>■ ≥ 0.32 Medium</font>  "
        "<font color='#991b1b'>■ &lt; 0.32 Low</font>  "
        "  |  Generated by FaceTrack AI · Narula Public School",
        style("footer", fontSize=6.5, fontName="Helvetica",
              textColor=MUTED, alignment=TA_CENTER)
    ))

    doc.build(elements)
    return str(path)


def _kpi_cell(label: str, value: str, colour):
    """Render a single KPI card cell for the summary row."""
    from reportlab.lib import colors
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.lib.enums import TA_CENTER
    from reportlab.platypus import Table, TableStyle, Paragraph, Spacer

    S_VAL = ParagraphStyle("kv", fontSize=16, fontName="Helvetica-Bold",
                           textColor=colour, alignment=TA_CENTER, leading=20)
    S_LBL = ParagraphStyle("kl", fontSize=7, fontName="Helvetica",
                           textColor=colors.HexColor("#64748b"),
                           alignment=TA_CENTER, leading=10)

    inner = Table(
        [[Paragraph(value, S_VAL)], [Paragraph(label.upper(), S_LBL)]],
        colWidths=["100%"]
    )
    inner.setStyle(TableStyle([
        ("BACKGROUND",   (0,0), (-1,-1), colors.HexColor("#f8fafc")),
        ("LINEBELOW",    (0,0), (-1,0),  0.8, colour),
        ("LEFTPADDING",  (0,0), (-1,-1), 6),
        ("RIGHTPADDING", (0,0), (-1,-1), 6),
        ("TOPPADDING",   (0,0), (0,0),   6),
        ("BOTTOMPADDING",(0,1), (-1,-1), 6),
        ("ROUNDEDCORNERS", [4]),
    ]))
    return inner


# ══════════════════════════════════════════════════════════════════════════════
# Excel
# ══════════════════════════════════════════════════════════════════════════════

def generate_excel_report(target_date: str, class_name: str = None,
                          stream: str = None, section: str = None) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import (
        PatternFill, Font, Alignment, Border, Side, numbers
    )
    from openpyxl.utils import get_column_letter

    fname = f"attendance_{target_date}"
    if class_name: fname += f"_{class_name}"
    if stream:     fname += f"_{stream}"
    if section:    fname += f"_Sec{section}"
    path = config.REPORT_DIR / (fname + ".xlsx")

    records = db.get_attendance_by_date(target_date, class_name, stream, section)
    # Deduplicate
    seen = {}
    for r in records:
        if r["student_id"] not in seen:
            seen[r["student_id"]] = r
    unique = sorted(
        seen.values(),
        key=lambda r: (r.get("class_name",""), r.get("stream",""),
                       r.get("section",""), r.get("roll_no",0))
    )

    all_students = db.list_students(class_name, stream, section)

    wb = Workbook()

    # ── Sheet 1: Attendance ────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Attendance"

    # Colour helpers
    def fill(hex_col):
        return PatternFill("solid", fgColor=hex_col.lstrip("#"))

    def border(style="thin"):
        s = Side(style=style, color="E2E8F0")
        return Border(left=s, right=s, top=s, bottom=s)

    BLUE_FILL   = fill("#1d4ed8")
    HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=10)
    BODY_FONT   = Font(name="Calibri", size=10)
    ALT_FILL    = fill("#F8FAFC")
    GREEN_FILL  = fill("#DCFCE7"); GREEN_FONT = Font(name="Calibri", color="166534", bold=True, size=10)
    AMBER_FILL  = fill("#FEF3C7"); AMBER_FONT = Font(name="Calibri", color="92400E", bold=True, size=10)
    RED_FILL    = fill("#FEE2E2"); RED_FONT   = Font(name="Calibri", color="991B1B", bold=True, size=10)
    CTR = Alignment(horizontal="center", vertical="center")
    LFT = Alignment(horizontal="left",   vertical="center")

    # ── Title block ────────────────────────────────────────────────────────────
    ws.merge_cells("A1:I1")
    ws["A1"] = config.SCHOOL_NAME
    ws["A1"].font = Font(name="Calibri", bold=True, size=14, color="0F172A")
    ws["A1"].alignment = LFT

    ws.merge_cells("A2:I2")
    subtitle = f"Attendance Report  ·  {target_date}"
    if class_name: subtitle += f"  ·  Class {class_name}"
    if stream:     subtitle += f"  ·  {stream}"
    if section:    subtitle += f"  ·  Section {section}"
    ws["A2"] = subtitle
    ws["A2"].font = Font(name="Calibri", size=10, color="64748B")
    ws["A2"].alignment = LFT
    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18

    # ── KPI row ────────────────────────────────────────────────────────────────
    total_e = len(all_students)
    pres_c  = len(unique)
    rate    = round(pres_c / total_e * 100, 1) if total_e else 0
    avg_cf  = sum(r.get("confidence") or 0 for r in unique) / len(unique) if unique else 0

    ws["A4"] = "Present"; ws["B4"] = pres_c
    ws["C4"] = "Absent";  ws["D4"] = total_e - pres_c
    ws["E4"] = "Rate";    ws["F4"] = f"{rate}%"
    ws["G4"] = "Avg Confidence"; ws["H4"] = round(avg_cf, 3)

    for col in "ABCDEFGH":
        cell = ws[f"{col}4"]
        cell.font = Font(name="Calibri", bold=(col in "ACEG"), size=10,
                         color="1d4ed8" if col in "ACEG" else "0F172A")
        cell.alignment = LFT
    ws.row_dimensions[4].height = 18

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = ["#", "Student ID", "Name", "Class", "Stream", "Section",
               "Roll No", "First Detected", "Confidence"]
    for col_i, h in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_i, value=h)
        cell.fill      = BLUE_FILL
        cell.font      = HEADER_FONT
        cell.alignment = CTR
        cell.border    = border()
    ws.row_dimensions[6].height = 20

    # ── Data rows ──────────────────────────────────────────────────────────────
    for row_i, r in enumerate(unique, 7):
        conf = r.get("confidence") or 0
        vals = [
            row_i - 6,
            r["student_id"],
            r["name"],
            r.get("class_name", ""),
            r.get("stream", "") or "—",
            r.get("section", ""),
            r.get("roll_no", ""),
            str(r.get("detected_at", ""))[-8:] if r.get("detected_at") else "—",
            round(conf, 3),
        ]
        use_alt = (row_i % 2 == 0)
        for col_i, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            cell.border    = border()
            cell.alignment = CTR if col_i != 3 else LFT
            if col_i == 9:   # confidence column
                if conf >= 0.45:   cell.fill = GREEN_FILL; cell.font = GREEN_FONT
                elif conf >= 0.32: cell.fill = AMBER_FILL; cell.font = AMBER_FONT
                else:              cell.fill = RED_FILL;   cell.font = RED_FONT
            else:
                cell.font = BODY_FONT
                cell.fill = ALT_FILL if use_alt else PatternFill()
        ws.row_dimensions[row_i].height = 18

    # Column widths
    for col_i, w in enumerate([5, 14, 26, 8, 14, 10, 9, 17, 14], 1):
        ws.column_dimensions[get_column_letter(col_i)].width = w

    ws.freeze_panes = "A7"

    # ── Sheet 2: Section Summary ───────────────────────────────────────────────
    ws2 = wb.create_sheet("Section Summary")

    ws2.merge_cells("A1:G1")
    ws2["A1"] = "Section-wise Attendance Summary  ·  " + target_date
    ws2["A1"].font = Font(name="Calibri", bold=True, size=12, color="0F172A")
    ws2["A1"].alignment = LFT
    ws2.row_dimensions[1].height = 22

    hdrs2 = ["Class", "Stream", "Section", "Present", "Total", "Absent", "Rate (%)"]
    for col_i, h in enumerate(hdrs2, 1):
        cell = ws2.cell(row=3, column=col_i, value=h)
        cell.fill = BLUE_FILL; cell.font = HEADER_FONT
        cell.alignment = CTR; cell.border = border()
    ws2.row_dimensions[3].height = 20

    # Build section groups
    from collections import defaultdict
    groups = defaultdict(lambda: {"present": 0, "total": 0})
    for r in unique:
        key = (r.get("class_name",""), r.get("stream",""), r.get("section",""))
        groups[key]["present"] += 1
    for s in all_students:
        key = (s["class_name"], s.get("stream",""), s["section"])
        groups[key]["total"] += 1

    for row_i, ((cls, strm, sec), cnt) in enumerate(sorted(groups.items()), 4):
        pres  = cnt["present"]
        total = cnt["total"]
        ab    = total - pres
        rate2 = round(pres / total * 100, 1) if total else 0
        vals2 = [cls, strm or "—", sec, pres, total, ab, rate2]
        use_alt = (row_i % 2 == 0)
        for col_i, v in enumerate(vals2, 1):
            cell = ws2.cell(row=row_i, column=col_i, value=v)
            cell.border = border(); cell.alignment = CTR
            if col_i == 7:  # rate
                if rate2 >= 75:   cell.fill = GREEN_FILL; cell.font = GREEN_FONT
                elif rate2 >= 50: cell.fill = AMBER_FILL; cell.font = AMBER_FONT
                else:             cell.fill = RED_FILL;   cell.font = RED_FONT
            else:
                cell.font = BODY_FONT
                cell.fill = ALT_FILL if use_alt else PatternFill()
        ws2.row_dimensions[row_i].height = 18

    for col_i, w in enumerate([10, 14, 10, 10, 10, 10, 12], 1):
        ws2.column_dimensions[get_column_letter(col_i)].width = w

    ws2.freeze_panes = "A4"

    wb.save(str(path))
    return str(path)
